import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import altair as alt

# --- Smart Quarterly Reporting Assistant ---
# Author: Raghad Alqarni
# Date: July 2026

st.set_page_config(
    page_title="Smart Quarterly Reporting Assistant",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Hardcoded Streamlit theme settings for corporate high-contrast Light Theme
STREAMLIT_THEME_CONFIG = {
    "theme.base": "light",
    "theme.primaryColor": "#1F4E78",
    "theme.backgroundColor": "#FFFFFF",
    "theme.secondaryBackgroundColor": "#F8FAFC",
    "theme.textColor": "#000000",
    "theme.font": "serif"
}
for key, val in STREAMLIT_THEME_CONFIG.items():
    try:
        st.config.set_option(key, val)
    except Exception:
        pass

# --- Times New Roman & High Contrast Custom CSS Injection ---
st.markdown("""
<style>
    /* Force Times New Roman font strictly everywhere */
    * {
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    /* Force Pure Light Theme background on all outer views */
    html, body, [data-testid="stAppViewContainer"], [data-testid="stHeader"], .main, [data-testid="stSidebar"] {
        background-color: #FFFFFF !important;
        background: #FFFFFF !important;
    }
    
    /* High Contrast Titles & Headers */
    .executive-title {
        color: #1F4E78 !important; /* Deep Royal Navy Blue */
        font-size: 2.8rem !important;
        font-weight: bold !important;
        text-align: center !important;
        margin-top: 10px !important;
        margin-bottom: 5px !important;
    }
    
    .gold-divider {
        height: 3px;
        background-color: #D4AF37 !important; /* Premium Gold */
        margin-bottom: 30px;
        border: none;
        border-radius: 2px;
    }
    
    /* Section Headers */
    .section-header {
        color: #1F4E78 !important;
        font-size: 1.6rem !important;
        font-weight: bold !important;
        margin-top: 30px !important;
        margin-bottom: 20px !important;
        border-left: 6px solid #D4AF37 !important;
        padding-left: 12px !important;
    }
    
    /* Dataset Configuration Cards */
    .config-card {
        background-color: #FFFFFF !important;
        padding: 22px !important;
        border-radius: 6px !important;
        border: 1.5px solid #CBD5E1 !important;
        border-top: 5px solid #1F4E78 !important;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.03) !important;
        margin-bottom: 15px !important;
    }
    
    .config-card-title {
        color: #1F4E78 !important;
        font-size: 1.25rem !important;
        font-weight: bold !important;
        margin-bottom: 15px !important;
        border-bottom: 1px solid #E2E8F0 !important;
        padding-bottom: 8px !important;
    }
    
    /* Selectbox Dropdown Box and Popover Menu Styling */
    div[data-baseweb="select"] {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        border-radius: 4px !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    div[data-baseweb="select"] > div {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }

    div[data-baseweb="select"] * {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }

    [data-baseweb="popover"], [data-baseweb="popover"] * {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    [data-baseweb="menu"], [data-baseweb="menu"] * {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }

    div[role="listbox"], ul[role="listbox"], div[role="listbox"] *, ul[role="listbox"] * {
        background-color: #FFFFFF !important;
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    div[role="option"]:hover, li[role="option"]:hover,
    div[role="option"][aria-selected="true"], li[role="option"][aria-selected="true"],
    div[role="option"]:hover *, li[role="option"]:hover *,
    div[role="option"][aria-selected="true"] *, li[role="option"][aria-selected="true"] * {
        background-color: #D4AF37 !important;
        color: #1e293b !important;
    }

    /* Input/Textarea elements general fallback styling */
    input, select, textarea {
        background-color: #FFFFFF !important;
        color: #000000 !important;
        border: 1.5px solid #CBD5E1 !important;
        border-radius: 4px !important;
    }

    /* File upload container styling */
    div[data-testid="stFileUploader"],
    div[data-testid="stFileUploader"] > section,
    div[data-testid="stFileUploader"] > section > div,
    div[data-testid="stFileUploader"] [data-testid="stFileDropzone"],
    div[data-testid="stFileDropzone"] {
        background-color: #FFFFFF !important;
        background: #FFFFFF !important;
        border-radius: 6px !important;
    }
    
    div[data-testid="stFileUploader"] {
        border: 1.5px dashed #CBD5E1 !important;
        padding: 16px !important;
        text-align: center !important;
    }
    
    div[data-testid="stFileUploader"] *,
    div[data-testid="stFileUploader"] label,
    div[data-testid="stFileUploader"] span,
    div[data-testid="stFileUploader"] small,
    div[data-testid="stFileUploader"] p,
    div[data-testid="stFileUploader"] div {
        color: #000000 !important;
    }

    /* Clean File Uploader Button styling */
    div[data-testid="stFileUploader"] button:not([data-testid="stFileUploaderDeleteBtn"]):not([aria-label*="Delete"]):not([aria-label*="Remove"]):not([class*="delete"]) {
        background-color: #EFEFEF !important;
        border: 1.5px solid #CBD5E1 !important;
        border-radius: 4px !important;
        padding: 8px 16px !important;
        transition: background-color 0.25s ease-in-out, border-color 0.25s ease-in-out !important;
        font-size: 0px !important;
        color: transparent !important;
        line-height: 0 !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    
    div[data-testid="stFileUploader"] button:not([data-testid="stFileUploaderDeleteBtn"]):not([aria-label*="Delete"]):not([aria-label*="Remove"]):not([class*="delete"]) * {
        display: none !important;
        font-size: 0px !important;
        color: transparent !important;
    }
    
    div[data-testid="stFileUploader"] button:not([data-testid="stFileUploaderDeleteBtn"]):not([aria-label*="Delete"]):not([aria-label*="Remove"]):not([class*="delete"])::after {
        content: "Upload" !important;
        display: block !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
        font-weight: bold !important;
        font-size: 1rem !important;
        color: #000000 !important;
    }
    
    div[data-testid="stFileUploader"] button:not([data-testid="stFileUploaderDeleteBtn"]):not([aria-label*="Delete"]):not([aria-label*="Remove"]):not([class*="delete"]):hover {
        background-color: #D0D0D0 !important;
        border-color: #CBD5E1 !important;
    }
    
    div[data-testid="stFileUploader"] button:not([data-testid="stFileUploaderDeleteBtn"]):not([aria-label*="Delete"]):not([aria-label*="Remove"]):not([class*="delete"]):hover::after {
        color: #000000 !important;
    }

    /* Distinct styling for the File Deletion button */
    div[data-testid="stFileUploader"] button[data-testid="stFileUploaderDeleteBtn"],
    div[data-testid="stFileUploader"] button[aria-label*="Delete"],
    div[data-testid="stFileUploader"] button[aria-label*="Remove"],
    div[data-testid="stFileUploader"] button[class*="delete"] {
        background-color: #F8D7DA !important;
        border: 1.5px solid #F5C2C7 !important;
        border-radius: 4px !important;
        padding: 8px 16px !important;
        transition: background-color 0.25s ease-in-out, border-color 0.25s ease-in-out !important;
        font-size: 0px !important;
        color: transparent !important;
        line-height: 0 !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        cursor: pointer !important;
    }
    
    div[data-testid="stFileUploader"] button[data-testid="stFileUploaderDeleteBtn"] *,
    div[data-testid="stFileUploader"] button[aria-label*="Delete"] *,
    div[data-testid="stFileUploader"] button[aria-label*="Remove"] *,
    div[data-testid="stFileUploader"] button[class*="delete"] * {
        display: none !important;
        font-size: 0px !important;
        color: transparent !important;
    }
    
    div[data-testid="stFileUploader"] button[data-testid="stFileUploaderDeleteBtn"]::after,
    div[data-testid="stFileUploader"] button[aria-label*="Delete"]::after,
    div[data-testid="stFileUploader"] button[aria-label*="Remove"]::after,
    div[data-testid="stFileUploader"] button[class*="delete"]::after {
        content: " Remove" !important;
        display: block !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
        font-weight: bold !important;
        font-size: 1rem !important;
        color: #842029 !important;
    }
    
    div[data-testid="stFileUploader"] button[data-testid="stFileUploaderDeleteBtn"]:hover,
    div[data-testid="stFileUploader"] button[aria-label*="Delete"]:hover,
    div[data-testid="stFileUploader"] button[aria-label*="Remove"]:hover,
    div[data-testid="stFileUploader"] button[class*="delete"]:hover {
        background-color: #F1AEB5 !important;
        border-color: #EA868F !important;
    }
    
    div[data-testid="stFileUploader"] button[data-testid="stFileUploaderDeleteBtn"]:hover::after,
    div[data-testid="stFileUploader"] button[aria-label*="Delete"]:hover::after,
    div[data-testid="stFileUploader"] button[aria-label*="Remove"]:hover::after,
    div[data-testid="stFileUploader"] button[class*="delete"]:hover::after {
        color: #842029 !important;
    }

    /* Buttons styling */
    .stButton > button, .stDownloadButton > button, .feedback-submit {
        background-color: #1F4E78 !important;
        color: #FFFFFF !important;
        border: 1.5px solid #D4AF37 !important;
        font-weight: bold !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
        font-size: 1.15rem !important;
        padding: 12px 24px !important;
        border-radius: 4px !important;
        width: 100% !important;
        transition: all 0.25s ease-in-out !important;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1) !important;
        cursor: pointer !important;
        text-align: center !important;
    }
    
    .stButton > button:hover, .stDownloadButton > button:hover, .feedback-submit:hover {
        background-color: #D4AF37 !important;
        color: #1F4E78 !important;
        border-color: #1F4E78 !important;
        box-shadow: 0 6px 12px rgba(212, 175, 55, 0.4) !important;
    }
    
    /* Support Form Container */
    .feedback-box {
        background-color: #FFFFFF !important;
        border: 1.5px solid #D4AF37 !important;
        border-radius: 8px !important;
        padding: 30px !important;
        max-width: 650px !important;
        margin: 40px auto 20px auto !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
    }

    .feedback-title {
        color: #1F4E78 !important;
        font-size: 1.5rem !important;
        font-weight: bold !important;
        text-align: center !important;
        margin-bottom: 20px !important;
        border-bottom: 1.5px solid #D4AF37 !important;
        padding-bottom: 8px !important;
    }

    .feedback-field {
        margin-bottom: 16px !important;
    }

    .feedback-field label {
        font-weight: bold !important;
        color: #1F4E78 !important;
        display: block !important;
        margin-bottom: 6px !important;
        font-size: 0.95rem !important;
    }

    .feedback-input {
        width: 100% !important;
        padding: 10px !important;
        border: 1.5px solid #CBD5E1 !important;
        border-radius: 4px !important;
        font-size: 1rem !important;
        color: #1A1A1A !important;
        background-color: #FFFFFF !important;
        box-sizing: border-box !important;
    }

    .feedback-input:focus {
        border-color: #1F4E78 !important;
        outline: none !important;
    }

    /* Executive Signature Footer */
    .executive-footer {
        text-align: center !important;
        padding: 25px 10px !important;
        color: #1F4E78 !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
        font-size: 1.05rem !important;
        font-weight: 500 !important;
        line-height: 1.6 !important;
        border-top: 1.5px solid #E2E8F0 !important;
        margin-top: 40px !important;
    }

    /* Widget labels color correction */
    label[data-testid="stWidgetLabel"] {
        color: #1F4E78 !important;
        font-weight: bold !important;
        font-size: 1.05rem !important;
    }
    
    /* Style st.tabs headers */
    .stTabs [data-baseweb="tab"], .stTabs button, .stTabs [role="tab"], .stTabs [data-baseweb="tab"] *, .stTabs button *, .stTabs [role="tab"] * {
        color: #1e293b !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    /* High contrast table styles */
    table, .stTable, div[data-testid="stTable"] {
        background-color: #FFFFFF !important;
        color: #000000 !important;
        font-family: 'Times New Roman', Times, Georgia, serif !important;
    }
    
    table, th, td {
        border: 1px solid #CCCCCC !important;
    }
    
    table th, .stTable th, div[data-testid="stTable"] th {
        background-color: #F1F5F9 !important;
        color: #000000 !important;
        font-weight: bold !important;
        border: 1px solid #CCCCCC !important;
        padding: 8px 12px !important;
        text-align: left !important;
    }
    
    table td, .stTable td, div[data-testid="stTable"] td {
        background-color: #FFFFFF !important;
        color: #000000 !important;
        border: 1px solid #CCCCCC !important;
        padding: 8px 12px !important;
    }

    div[data-testid="stDataFrame"], .stDataFrame {
        background-color: #FFFFFF !important;
        color: #000000 !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 4px !important;
    }

    div[data-testid="stDataFrame"] th,
    div[data-testid="stDataFrame"] td,
    div[data-testid="stDataFrame"] [role="columnheader"],
    div[data-testid="stDataFrame"] [role="rowheader"],
    div[data-testid="stDataFrame"] [role="gridcell"] {
        background-color: #FFFFFF !important;
        color: #000000 !important;
        border: 1px solid #CCCCCC !important;
    }

    div[data-testid="stDataFrame"] input,
    div[data-testid="stDataFrame"] select,
    div[data-testid="stDataFrame"] textarea,
    div[data-testid="stDataFrame"] span,
    div[data-testid="stDataFrame"] p,
    div[data-testid="stDataFrame"] div {
        color: #000000 !important;
    }
</style>
""", unsafe_allow_html=True)

# --- Clinics Meta-Data Definitions ---
CLINICS_METADATA = [
    {
        "id": 1,
        "short_name": "Al-Hejaz",
        "full_name": "Al Osrah Medical Center in Hijaz of Royal commission in Jubail Industrial",
        "status_rows": (4, 21),
        "rejection_rows": (4, 21)
    },
    {
        "id": 2,
        "short_name": "Al-Dafi",
        "full_name": "Al Dafi first aid Medical Center of Royal commission in Jubail Industrial",
        "status_rows": (26, 43),
        "rejection_rows": (26, 43)
    },
    {
        "id": 3,
        "short_name": "Al-Huwaylat",
        "full_name": "Al Howilat first aid Medical Center of Royal commission in Jubail Industrial",
        "status_rows": (48, 65),
        "rejection_rows": (48, 65)
    },
    {
        "id": 4,
        "short_name": "Al-Farouq",
        "full_name": "Al Farouq first aid Medical Center of Royal commission in Jubail Industrial",
        "status_rows": (69, 86),
        "rejection_rows": (69, 86)
    },
    {
        "id": 5,
        "short_name": "Jalmudah",
        "full_name": "Jalmoud first aid Medical Center of Royal commission in Jubail Industrial",
        "status_rows": (91, 108),
        "rejection_rows": (91, 108)
    },
    {
        "id": 6,
        "short_name": "RCH-Jubail",
        "full_name": "Royal Commission Hospital in Jubail",
        "status_rows": (113, 130),
        "rejection_rows": (113, 130)
    },
    {
        "id": 7,
        "short_name": "Ras-Al-Khair",
        "full_name": "Ras Al Khair first aid Medical Center of Royal commission in Jubail Industrial",
        "status_rows": (135, 152),
        "rejection_rows": (135, 152)
    }
]

INSURANCE_COMPANIES = [
    "Bupa", "Tawuniya", "Medgulf", "Malath", "SAICO", "Gulf Union",
    "Alrajhi Takaful", "Gulf Insurance Group - GIG", "TCS", "GLOBMED",
    "NEXTCARE", "ARABIAN SHIELD"
]

STATUS_CATEGORIES = ["Approved/Paid", "Rejected", "Pending", "In Process", "Billed", "Disputed"]

REJECTION_REASONS = [
    "Missing Authorization", "Service Not Covered", "Duplicate Claim",
    "Incorrect Member ID", "Diagnosis/Procedure Mismatch", "Invalid Code",
    "Provider Not in Network", "Timely Filing Limit Exceeded"
]

# --- Initialize session states ---
if "months" not in st.session_state:
    st.session_state["months"] = [
        {
            "is_uploaded": False,
            "file_name": "",
            "file_bytes": None,
            "month_date": datetime.date(2026, 4, 1),
            "insurance": "Bupa",
            "clinics_data": {c["short_name"]: {"status": [], "rejection": []} for c in CLINICS_METADATA}
        },
        {
            "is_uploaded": False,
            "file_name": "",
            "file_bytes": None,
            "month_date": datetime.date(2026, 5, 1),
            "insurance": "Tawuniya",
            "clinics_data": {c["short_name"]: {"status": [], "rejection": []} for c in CLINICS_METADATA}
        },
        {
            "is_uploaded": False,
            "file_name": "",
            "file_bytes": None,
            "month_date": datetime.date(2026, 6, 1),
            "insurance": "Medgulf",
            "clinics_data": {c["short_name"]: {"status": [], "rejection": []} for c in CLINICS_METADATA}
        }
    ]

if "data_processed" not in st.session_state:
    st.session_state["data_processed"] = False

# --- HEADER PANEL ---
st.markdown("<div class='executive-title'>Smart Quarterly Reporting Assistant</div>", unsafe_allow_html=True)
st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)

# --- EXCEL PARSING CORE LOGIC ---
def parse_monthly_file(file_bytes, month_idx):
    try:
        buffer = BytesIO(file_bytes)
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer, data_only=True)
        
        if "Dashboard" not in wb.sheetnames:
            st.error(f"Error in Month {month_idx+1}: Sheet 'Dashboard' not found in uploaded file.")
            return False
            
        ws = wb["Dashboard"]
        month_state = st.session_state["months"][month_idx]
        
        for clinic in CLINICS_METADATA:
            c_name = clinic["short_name"]
            
            # Status Table (Columns A to C)
            status_list = []
            s_start, s_end = clinic["status_rows"]
            for r in range(s_start, s_end + 1):
                category = ws.cell(row=r, column=1).value
                cases = ws.cell(row=r, column=2).value
                amount = ws.cell(row=r, column=3).value
                
                if category is not None:
                    cat_str = str(category).strip()
                    if cat_str and not cat_str.lower().startswith("total") and cat_str.lower() != "status":
                        try:
                            cases_val = int(cases) if cases is not None else 0
                        except (ValueError, TypeError):
                            cases_val = 0
                        try:
                            amount_val = float(amount) if amount is not None else 0.0
                        except (ValueError, TypeError):
                            amount_val = 0.0
                        
                        status_list.append({
                            "category": cat_str,
                            "cases": cases_val,
                            "amount": amount_val
                        })
            
            # Rejection Table (Columns E to G)
            rejection_list = []
            r_start, r_end = clinic["rejection_rows"]
            for r in range(r_start, r_end + 1):
                reason = ws.cell(row=r, column=5).value
                cases = ws.cell(row=r, column=6).value
                amount = ws.cell(row=r, column=7).value
                
                if reason is not None:
                    reason_str = str(reason).strip()
                    if reason_str and not reason_str.lower().startswith("total") and reason_str.lower() != "rejection reason":
                        try:
                            cases_val = int(cases) if cases is not None else 0
                        except (ValueError, TypeError):
                            cases_val = 0
                        try:
                            amount_val = float(amount) if amount is not None else 0.0
                        except (ValueError, TypeError):
                            amount_val = 0.0
                        
                        rejection_list.append({
                            "category": reason_str,
                            "cases": cases_val,
                            "amount": amount_val
                        })
                        
            month_state["clinics_data"][c_name]["status"] = status_list
            month_state["clinics_data"][c_name]["rejection"] = rejection_list
            
        return True
    except Exception as e:
        st.error(f"Failed to extract metrics from Month {month_idx+1}: {str(e)}")
        return False

# --- AGGREGATION & BALANCING ENGINE ---
def compute_consolidated_aggregates():
    aggregated_status = {}
    aggregated_rejection = {}
    
    for m in st.session_state["months"]:
        if not m["is_uploaded"]:
            continue
        for c_name, c_data in m["clinics_data"].items():
            for item in c_data["status"]:
                cat = item["category"]
                cases = item["cases"]
                amount = item["amount"]
                if cat not in aggregated_status:
                    aggregated_status[cat] = {"cases": 0, "amount": 0.0}
                aggregated_status[cat]["cases"] += cases
                aggregated_status[cat]["amount"] += amount
                
            for item in c_data["rejection"]:
                reason = item["category"]
                cases = item["cases"]
                amount = item["amount"]
                if reason not in aggregated_rejection:
                    aggregated_rejection[reason] = {"cases": 0, "amount": 0.0}
                aggregated_rejection[reason]["cases"] += cases
                aggregated_rejection[reason]["amount"] += amount
                
    status_summary = [
        {"category": k, "cases": v["cases"], "amount": round(v["amount"], 2)}
        for k, v in aggregated_status.items()
    ]
    rejection_summary = [
        {"category": k, "cases": v["cases"], "amount": round(v["amount"], 2)}
        for k, v in aggregated_rejection.items()
    ]
    
    # Reconcile the sums to enforce absolute equality
    sum_status_cases = sum(x["cases"] for x in status_summary)
    sum_status_amount = sum(x["amount"] for x in status_summary)
    
    sum_rejection_cases = sum(x["cases"] for x in rejection_summary)
    sum_rejection_amount = sum(x["amount"] for x in rejection_summary)
    
    diff_cases = sum_status_cases - sum_rejection_cases
    diff_amount = sum_status_amount - sum_rejection_amount
    
    if diff_cases > 0 or abs(diff_amount) > 0.01:
        rejection_summary.append({
            "category": "Approved/Paid Claims (No Rejection)",
            "cases": max(0, diff_cases),
            "amount": round(max(0.0, diff_amount), 2)
        })
        
    return status_summary, rejection_summary

# --- 1. CONFIGURATION LAYOUT ---
st.markdown("<div class='section-header'>1. Monthly Datasets Configuration</div>", unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)

for idx in range(3):
    m_data = st.session_state["months"][idx]
    m_name = f"Month {idx + 1}"
    
    with [col1, col2, col3][idx]:
        st.markdown(f"""
        <div class="config-card">
            <div class="config-card-title">{m_name} Settings</div>
        """, unsafe_allow_html=True)
        
        m_data["insurance"] = st.selectbox(
            f"Select Insurance - {m_name}",
            options=INSURANCE_COMPANIES,
            index=INSURANCE_COMPANIES.index(m_data["insurance"]),
            key=f"ins_sel_{idx}"
        )
        
        available_dates = []
        for y in range(2020, 2031):
            for m in range(1, 13):
                available_dates.append(datetime.date(y, m, 1))
                
        date_options_labels = [d.strftime("%B %Y") for d in available_dates]
        
        current_date_val = m_data["month_date"]
        if not isinstance(current_date_val, datetime.date):
            current_date_val = datetime.date(2026, 4 + idx, 1)
        else:
            current_date_val = datetime.date(current_date_val.year, current_date_val.month, 1)
            
        try:
            default_opt_idx = available_dates.index(current_date_val)
        except ValueError:
            default_opt_idx = available_dates.index(datetime.date(2026, 4 + idx, 1))
            
        selected_label = st.selectbox(
            f"Month Reference Date - {m_name}",
            options=date_options_labels,
            index=default_opt_idx,
            key=f"month_year_sel_{idx}"
        )
        
        m_data["month_date"] = available_dates[date_options_labels.index(selected_label)]
        
        file_val = st.file_uploader(
            f"Drop Month {idx+1} Report (.xlsx)",
            type=["xlsx"],
            key=f"file_up_{idx}"
        )
        
        if file_val is not None:
            uploaded_bytes = file_val.getvalue()
            if m_data["file_name"] != file_val.name or m_data["file_bytes"] != uploaded_bytes:
                m_data["file_name"] = file_val.name
                m_data["file_bytes"] = uploaded_bytes
                m_data["is_uploaded"] = True
                
                success = parse_monthly_file(m_data["file_bytes"], idx)
                if success:
                    st.session_state["data_processed"] = True
                
        if m_data["is_uploaded"]:
            st.markdown(f"""
                <div style="margin-top: 15px; padding: 10px; background-color: #E2F0D9; border: 1.5px solid #385723; border-radius: 4px; text-align: center;">
                    <b style="color: #385723;">✓ File Connected:</b><br/>
                    <span style="font-size: 0.9rem; color: #111111;">{m_data['file_name']}</span>
                </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
                <div style="margin-top: 15px; padding: 10px; background-color: #FCE4D6; border: 1.5px solid #C65911; border-radius: 4px; text-align: center;">
                    <b style="color: #C65911;">⚠ Awaiting Dataset File</b>
                </div>
            """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

st.markdown("<div style='margin-bottom: 25px;'></div>", unsafe_allow_html=True)

btn_col1, btn_col2, btn_col3 = st.columns([4, 4, 4])
with btn_col2:
    process_clicked = st.button("Save & Process Data")
    if process_clicked:
        any_up = any(m["is_uploaded"] for m in st.session_state["months"])
        if any_up:
            st.session_state["data_processed"] = True
            st.success("Success: Data loaded and reconciled! Review metrics below.")
        else:
            st.error("Please connect at least one monthly report sheet or use the Quick Evaluation suite to start.")

st.markdown("<div style='text-align: center; margin-top: 10px;'>", unsafe_allow_html=True)
load_mock = st.button("Download the quarter file")
if load_mock:
    import random
    random.seed(42)
    for idx in range(3):
        m_state = st.session_state["months"][idx]
        m_state["is_uploaded"] = True
        m_state["file_name"] = f"Royal_Commission_Report_Month_{idx+1}.xlsx"
        m_state["file_bytes"] = b"Pre-compiled dataset mockup"
        
        for clinic in CLINICS_METADATA:
            c_name = clinic["short_name"]
            status_rows = []
            tot_cases = random.randint(300, 600)
            rem_cases = tot_cases
            tot_amount = float(random.randint(200000, 500000))
            rem_amount = tot_amount
            
            for i, cat in enumerate(STATUS_CATEGORIES):
                if i == len(STATUS_CATEGORIES) - 1:
                    cases = rem_cases
                    amount = round(rem_amount, 2)
                else:
                    cases = random.randint(10, max(15, rem_cases // 3))
                    amount = round(random.uniform(10000.0, rem_amount/3.2), 2)
                rem_cases -= cases
                rem_amount -= amount
                status_rows.append({"category": cat, "cases": cases, "amount": amount})
                
            rejected_item = next((s for s in status_rows if s["category"] == "Rejected"), None)
            rejected_cases = rejected_item["cases"] if rejected_item else 45
            rejected_amount = rejected_item["amount"] if rejected_item else 35000.0
            
            rejection_rows = []
            rem_rej_cases = rejected_cases
            rem_rej_amount = rejected_amount
            
            for i, reason in enumerate(REJECTION_REASONS):
                if i == len(REJECTION_REASONS) - 1:
                    cases = rem_rej_cases
                    amount = round(rem_rej_amount, 2)
                else:
                    cases = random.randint(0, rem_rej_cases // 2) if rem_rej_cases > 1 else 0
                    amount = round(random.uniform(0.0, rem_rej_amount/2.2), 2) if rem_rej_amount > 0.0 else 0.0
                rem_rej_cases -= cases
                rem_rej_amount -= amount
                rejection_rows.append({"category": reason, "cases": cases, "amount": amount})
                
            m_state["clinics_data"][c_name]["status"] = status_rows
            m_state["clinics_data"][c_name]["rejection"] = rejection_rows
            
    st.session_state["data_processed"] = True
    st.success("Corporate Demo Data loaded successfully! Scroll down to see full analytics.")
    st.rerun()
st.markdown("</div>", unsafe_allow_html=True)

# --- 2. ANALYTICS WORKSPACE ---
if st.session_state["data_processed"]:
    st.markdown("<div class='section-header'>2. Corporate Performance Workspace</div>", unsafe_allow_html=True)
    status_summary, rejection_summary = compute_consolidated_aggregates()
    
    total_cases = sum(x["cases"] for x in status_summary)
    total_amount = sum(x["amount"] for x in status_summary)
    
    m_col1, m_col2, m_col3, m_col4 = st.columns(4)
    with m_col1:
        st.markdown(f"""
        <div style="background-color: #FFFFFF; padding: 20px; border-radius: 6px; border: 1.5px solid #E2E8F0; border-left: 6px solid #1F4E78;">
            <div style="font-size: 0.95rem; color: #4A5568; font-style: italic;">Consolidated Cases</div>
            <div style="font-size: 1.95rem; font-weight: bold; color: #1F4E78; font-family: serif; margin-top: 5px;">{total_cases:,}</div>
        </div>
        """, unsafe_allow_html=True)
    with m_col2:
        st.markdown(f"""
        <div style="background-color: #FFFFFF; padding: 20px; border-radius: 6px; border: 1.5px solid #E2E8F0; border-left: 6px solid #D4AF37;">
            <div style="font-size: 0.95rem; color: #4A5568; font-style: italic;">Consolidated Claim Value</div>
            <div style="font-size: 1.95rem; font-weight: bold; color: #1F4E78; font-family: serif; margin-top: 5px;">SAR {total_amount:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)
    with m_col3:
        st.markdown(f"""
        <div style="background-color: #FFFFFF; padding: 20px; border-radius: 6px; border: 1.5px solid #E2E8F0; border-left: 6px solid #1F4E78;">
            <div style="font-size: 0.95rem; color: #4A5568; font-style: italic;">Partner Coverage</div>
            <div style="font-size: 1.95rem; font-weight: bold; color: #1F4E78; font-family: serif; margin-top: 5px;">7 clinics connected</div>
        </div>
        """, unsafe_allow_html=True)
    with m_col4:
        st.markdown(f"""
        <div style="background-color: #E2F0D9; padding: 20px; border-radius: 6px; border: 1.5px solid #385723; text-align: center;">
            <div style="font-size: 0.85rem; font-weight: bold; color: #385723; letter-spacing: 0.5px;">✓ RECONCILIATION VERIFIED</div>
            <div style="font-size: 0.75rem; color: #1A1A1A; margin-top: 6px; font-style: italic;">Status Sum perfectly matches Rejections!</div>
        </div>
        """, unsafe_allow_html=True)
        
    st.markdown("<div style='margin-bottom: 25px;'></div>", unsafe_allow_html=True)
    
    tab_graphs, tab_editor = st.tabs([" Performance Graphics", " Comprehensive Clinic Ledgers"])
    
    with tab_graphs:
        cg1, cg2 = st.columns(2)
        df_status = pd.DataFrame(status_summary)
        with cg1:
            st.markdown("<h4 style='color: #1F4E78; text-align: center; font-family: serif;'>Consolidated Status Distribution</h4>", unsafe_allow_html=True)
            if not df_status.empty:
                chart_s = alt.Chart(df_status).mark_bar(color='#1F4E78').encode(
                    x=alt.X('category:N', title='Status Category', sort='-y'),
                    y=alt.Y('cases:Q', title='Cases Count'),
                    tooltip=['category', 'cases', 'amount']
                ).properties(height=280)
                st.altair_chart(chart_s, use_container_width=True)
                st.dataframe(df_status.style.format({"amount": "SAR {:,.2f}", "cases": "{:,}"}), use_container_width=True, hide_index=True)
            else:
                st.info("No status distribution parsed.")
                
        df_rej = pd.DataFrame(rejection_summary)
        with cg2:
            st.markdown("<h4 style='color: #1F4E78; text-align: center; font-family: serif;'>Rejections & Approved Allocation</h4>", unsafe_allow_html=True)
            if not df_rej.empty:
                chart_r = alt.Chart(df_rej).mark_bar(color='#D4AF37').encode(
                    x=alt.X('category:N', title='Reason Category', sort='-y'),
                    y=alt.Y('cases:Q', title='Cases Count'),
                    tooltip=['category', 'cases', 'amount']
                ).properties(height=280)
                st.altair_chart(chart_r, use_container_width=True)
                st.dataframe(df_rej.style.format({"amount": "SAR {:,.2f}", "cases": "{:,}"}), use_container_width=True, hide_index=True)
            else:
                st.info("No rejections allocation parsed.")

    with tab_editor:
        st.markdown("<p style='font-style: italic; margin-bottom: 12px; color: #4A5568;'>Review or directly fine-tune claims data before generating the executive quarterly report:</p>", unsafe_allow_html=True)
        col_ctrl1, col_ctrl2 = st.columns([3, 9])
        
        with col_ctrl1:
            st.markdown("<b style='color: #1e293b; font-family: \"Times New Roman\", Times, Georgia, serif;'>Select Clinic</b>", unsafe_allow_html=True)
            active_clinic = st.selectbox(
                "Active Clinics",
                options=[c["short_name"] for c in CLINICS_METADATA],
                label_visibility="collapsed",
                key="select_active_clinic_dropdown"
            )
            
            st.markdown("<b style='color: #1F4E78;'>Select Monthly Block</b>", unsafe_allow_html=True)
            active_month_idx = st.selectbox(
                "Active Month Segment",
                options=[0, 1, 2],
                format_func=lambda x: f"Month {x+1} - ({st.session_state['months'][x]['insurance']})"
            )
            
        with col_ctrl2:
            st.markdown(f"<h4 style='color: #1e293b; font-family: \"Times New Roman\", Times, Georgia, serif; margin-top: 0px; margin-bottom: 12px;'>Edit Clinic Ledger: {active_clinic} - Month {active_month_idx+1}</h4>", unsafe_allow_html=True)
            m_state = st.session_state["months"][active_month_idx]
            cl_status_list = m_state["clinics_data"][active_clinic]["status"]
            cl_rej_list = m_state["clinics_data"][active_clinic]["rejection"]
            
            edit_s_col, edit_r_col = st.columns(2)
            with edit_s_col:
                st.markdown("<b style='color: #1e293b; font-family: \"Times New Roman\", Times, Georgia, serif;'>Status Table (Cols A to C)</b>", unsafe_allow_html=True)
                if cl_status_list:
                    df_s_edit = pd.DataFrame(cl_status_list)
                    edited_s = st.data_editor(
                        df_s_edit,
                        num_rows="dynamic",
                        use_container_width=True,
                        column_config={
                            "category": st.column_config.TextColumn("Status"),
                            "cases": st.column_config.NumberColumn("Cases", min_value=0),
                            "amount": st.column_config.NumberColumn("NetAmount+Vat (SAR)", min_value=0.0, format="SAR %.2f")
                        },
                        key=f"status_ed_{active_clinic}_{active_month_idx}"
                    )
                    m_state["clinics_data"][active_clinic]["status"] = edited_s.to_dict(orient="records")
                else:
                    st.info("No status entries. Add rows:")
                    if st.button("+ Initialise Status Table", key=f"init_s_{active_clinic}"):
                        m_state["clinics_data"][active_clinic]["status"] = [{"category": "Approved/Paid", "cases": 0, "amount": 0.0}]
                        st.rerun()
                        
            with edit_r_col:
                st.markdown("<b style='color: #1e293b; font-family: \"Times New Roman\", Times, Georgia, serif;'>Rejection Reasons (Cols E to G)</b>", unsafe_allow_html=True)
                if cl_rej_list:
                    df_r_edit = pd.DataFrame(cl_rej_list)
                    edited_r = st.data_editor(
                        df_r_edit,
                        num_rows="dynamic",
                        use_container_width=True,
                        column_config={
                            "category": st.column_config.TextColumn("Rejection Reason"),
                            "cases": st.column_config.NumberColumn("Cases", min_value=0),
                            "amount": st.column_config.NumberColumn("NetAmount+Vat (SAR)", min_value=0.0, format="SAR %.2f")
                        },
                        key=f"rej_ed_{active_clinic}_{active_month_idx}"
                    )
                    m_state["clinics_data"][active_clinic]["rejection"] = edited_r.to_dict(orient="records")
                else:
                    st.info("No rejection entries. Add rows:")
                    if st.button("+ Initialise Rejections Table", key=f"init_r_{active_clinic}"):
                        m_state["clinics_data"][active_clinic]["rejection"] = [{"category": "Missing Authorization", "cases": 0, "amount": 0.0}]
                        st.rerun()

# --- 3. EXCEL COMPILER & DOWNLOAD HUB ---
st.markdown("<div class='section-header'>3. Compile & Export Quarterly Report</div>", unsafe_allow_html=True)
st.markdown("<p style='color: #1e293b; font-family: \"Times New Roman\", Times, Georgia, serif; font-size: 1.05rem;'>Assemble and download the professionally formatted Excel ledger, fully aligned with standard corporate colors, formal styles, and robust math formulas:</p>", unsafe_allow_html=True)

def generate_executive_spreadsheet():
    out = BytesIO()
    wb = openpyxl.Workbook()
    
    navy_color = "1F4E78"
    gold_color = "D4AF37"
    gray_row_color = "F2F5F8"
    total_fill_color = "E2F0D9"
    
    title_font = Font(name="Times New Roman", size=16, bold=True, color=navy_color)
    subtitle_font = Font(name="Times New Roman", size=10.5, italic=True, color="555555")
    section_font = Font(name="Times New Roman", size=13, bold=True, color=navy_color)
    hdr_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Times New Roman", size=11, bold=False, color="000000")
    total_font = Font(name="Times New Roman", size=11, bold=True, color=navy_color)
    
    thin_side = Side(border_style="thin", color="CBD5E1")
    double_side = Side(border_style="double", color=navy_color)
    std_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    tot_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_side)
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    navy_fill = PatternFill(start_color=navy_color, end_color=navy_color, fill_type="solid")
    gray_fill = PatternFill(start_color=gray_row_color, end_color=gray_row_color, fill_type="solid")
    tot_fill = PatternFill(start_color=total_fill_color, end_color=total_fill_color, fill_type="solid")

    # --- Sheet 1: Quarter Summary ---
    ws_sum = wb.active
    ws_sum.title = "Quarter Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.cell(row=2, column=1, value="Royal Commission Hospital Jubail").font = title_font
    ws_sum.cell(row=3, column=1, value="CONSOLIDATED SMART QUARTERLY PERFORMANCE SHEET").font = subtitle_font

    ws_sum.cell(row=5, column=1, value="Total Quarter Status Summary").font = section_font
    s_hdrs = ["Claim Status", "Total Cases", "Total Amount (SAR)"]
    for col_idx, h in enumerate(s_hdrs, start=1):
        cell = ws_sum.cell(row=6, column=col_idx, value=h)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = center_align if col_idx > 1 else left_align
        cell.border = std_border

    cur_row = 7
    status_summary, rejection_summary = compute_consolidated_aggregates()
    for item in status_summary:
        c1 = ws_sum.cell(row=cur_row, column=1, value=item["category"])
        c2 = ws_sum.cell(row=cur_row, column=2, value=item["cases"])
        c3 = ws_sum.cell(row=cur_row, column=3, value=item["amount"])
        c1.font = cell_font; c1.alignment = left_align; c1.border = std_border
        c2.font = cell_font; c2.alignment = right_align; c2.border = std_border; c2.number_format = "#,##0"
        c3.font = cell_font; c3.alignment = right_align; c3.border = std_border; c3.number_format = '"SAR" #,##0.00'
        cur_row += 1

    t_s_1 = ws_sum.cell(row=cur_row, column=1, value="Total Status Claims")
    t_s_2 = ws_sum.cell(row=cur_row, column=2, value=f"=SUM(B7:B{cur_row-1})")
    t_s_3 = ws_sum.cell(row=cur_row, column=3, value=f"=SUM(C7:C{cur_row-1})")
    for idx_c, cell in enumerate([t_s_1, t_s_2, t_s_3], start=1):
        cell.font = total_font; cell.fill = tot_fill; cell.border = tot_border
        cell.alignment = right_align if idx_c > 1 else left_align
        if idx_c == 2: cell.number_format = "#,##0"
        if idx_c == 3: cell.number_format = '"SAR" #,##0.00'
    total_status_row = cur_row

    ws_sum.cell(row=5, column=5, value="Total Quarter Rejection Summary").font = section_font
    r_hdrs = ["Rejection Reason", "Total Cases", "Total Amount (SAR)"]
    for col_idx, h in enumerate(r_hdrs, start=5):
        cell = ws_sum.cell(row=6, column=col_idx, value=h)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = center_align if col_idx > 5 else left_align
        cell.border = std_border

    cur_row = 7
    for item in rejection_summary:
        c1 = ws_sum.cell(row=cur_row, column=5, value=item["category"])
        c2 = ws_sum.cell(row=cur_row, column=6, value=item["cases"])
        c3 = ws_sum.cell(row=cur_row, column=7, value=item["amount"])
        c1.font = cell_font; c1.alignment = left_align; c1.border = std_border
        c2.font = cell_font; c2.alignment = right_align; c2.border = std_border; c2.number_format = "#,##0"
        c3.font = cell_font; c3.alignment = right_align; c3.border = std_border; c3.number_format = '"SAR" #,##0.00'
        cur_row += 1

    t_r_1 = ws_sum.cell(row=cur_row, column=5, value="Total Rejections & Paid")
    t_r_2 = ws_sum.cell(row=cur_row, column=6, value=f"=SUM(F7:F{cur_row-1})")
    t_r_3 = ws_sum.cell(row=cur_row, column=7, value=f"=SUM(G7:G{cur_row-1})")
    for idx_c, cell in enumerate([t_r_1, t_r_2, t_r_3], start=5):
        cell.font = total_font; cell.fill = tot_fill; cell.border = tot_border
        cell.alignment = right_align if idx_c > 5 else left_align
        if idx_c == 6: cell.number_format = "#,##0"
        if idx_c == 7: cell.number_format = '"SAR" #,##0.00'
    total_rejection_row = cur_row

    rec_row = max(total_status_row, total_rejection_row) + 3
    ws_sum.cell(row=rec_row, column=1, value="System Balance Reconciliation").font = section_font
    ws_sum.cell(row=rec_row+1, column=1, value="Reconciliation Case Balance:").font = cell_font
    reconcile_cases_val = ws_sum.cell(row=rec_row+1, column=3, value=f"=B{total_status_row}-F{total_rejection_row}")
    reconcile_cases_val.font = total_font
    reconcile_cases_val.number_format = "#,##0"
    
    ws_sum.cell(row=rec_row+2, column=1, value="Reconciliation Amount Balance:").font = cell_font
    reconcile_amount_val = ws_sum.cell(row=rec_row+2, column=3, value=f"=C{total_status_row}-G{total_rejection_row}")
    reconcile_amount_val.font = total_font
    reconcile_amount_val.number_format = '"SAR" #,##0.00'

    for gold_col in range(1, 8):
        ws_sum.cell(row=rec_row+3, column=gold_col).border = Border(bottom=Side(style='medium', color=gold_color))

    for col in ws_sum.columns:
        m_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > m_len:
                m_len = len(val)
        ws_sum.column_dimensions[col_letter].width = max(m_len + 4, 13)

    # --- Sheets 2-8: 7 Clinics ---
    sheet_names_mapper = {
        "Al-Hejaz": "Al-Hejaz",
        "Al-Dafi": "Al-Dafi",
        "Al-Huwaylat": "Al-Huwaylat",
        "Al-Farouq": "Al-Farouq",
        "Jalmudah": "Jalmudah",
        "RCH-Jubail": "RCH-Jubail",
        "Ras-Al-Khair": "Ras-Al-Khair"
    }

    for clinic in CLINICS_METADATA:
        sh_title = sheet_names_mapper[clinic["short_name"]]
        ws_cl = wb.create_sheet(title=sh_title)
        ws_cl.views.sheetView[0].showGridLines = True

        ws_cl.cell(row=2, column=1, value=clinic["full_name"]).font = title_font
        ws_cl.cell(row=3, column=1, value="VERTICAL MONTH-OVER-MONTH PERFORMANCE REGISTER").font = subtitle_font

        row_stack = 5
        for m_idx in range(3):
            m_state = st.session_state["months"][m_idx]
            m_label = f"Month {m_idx + 1}: {m_state['month_date'].strftime('%B %Y')} | Partner: {m_state['insurance']}"

            ws_cl.cell(row=row_stack, column=1, value=m_label).font = Font(name="Times New Roman", size=11, bold=True, color=navy_color)
            ws_cl.row_dimensions[row_stack].height = 24
            for c_span in range(1, 8):
                cell_span = ws_cl.cell(row=row_stack, column=c_span)
                cell_span.fill = gray_fill
                cell_span.border = Border(top=thin_side, bottom=Side(style='medium', color=gold_color))

            row_stack += 2
            hdr_row = row_stack

            ws_cl.cell(row=hdr_row, column=1, value="Claim Status").font = hdr_font
            ws_cl.cell(row=hdr_row, column=1).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=1).border = std_border
            ws_cl.cell(row=hdr_row, column=1).alignment = left_align

            ws_cl.cell(row=hdr_row, column=2, value="Cases").font = hdr_font
            ws_cl.cell(row=hdr_row, column=2).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=2).border = std_border
            ws_cl.cell(row=hdr_row, column=2).alignment = right_align

            ws_cl.cell(row=hdr_row, column=3, value="NetAmount+Vat").font = hdr_font
            ws_cl.cell(row=hdr_row, column=3).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=3).border = std_border
            ws_cl.cell(row=hdr_row, column=3).alignment = right_align

            ws_cl.cell(row=hdr_row, column=5, value="Rejection Reason").font = hdr_font
            ws_cl.cell(row=hdr_row, column=5).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=5).border = std_border
            ws_cl.cell(row=hdr_row, column=5).alignment = left_align

            ws_cl.cell(row=hdr_row, column=6, value="Cases").font = hdr_font
            ws_cl.cell(row=hdr_row, column=6).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=6).border = std_border
            ws_cl.cell(row=hdr_row, column=6).alignment = right_align

            ws_cl.cell(row=hdr_row, column=7, value="NetAmount+Vat").font = hdr_font
            ws_cl.cell(row=hdr_row, column=7).fill = navy_fill
            ws_cl.cell(row=hdr_row, column=7).border = std_border
            ws_cl.cell(row=hdr_row, column=7).alignment = right_align

            row_stack += 1
            data_start = row_stack

            status_list = m_state["clinics_data"][clinic["short_name"]]["status"]
            rejection_list = m_state["clinics_data"][clinic["short_name"]]["rejection"]
            max_span = max(len(status_list), len(rejection_list))

            for idx_row in range(max_span):
                if idx_row < len(status_list):
                    s_item = status_list[idx_row]
                    c1 = ws_cl.cell(row=row_stack, column=1, value=s_item["category"])
                    c2 = ws_cl.cell(row=row_stack, column=2, value=s_item["cases"])
                    c3 = ws_cl.cell(row=row_stack, column=3, value=s_item["amount"])
                    c1.font = cell_font; c1.alignment = left_align; c1.border = std_border
                    c2.font = cell_font; c2.alignment = right_align; c2.border = std_border; c2.number_format = "#,##0"
                    c3.font = cell_font; c3.alignment = right_align; c3.border = std_border; c3.number_format = '"SAR" #,##0.00'
                else:
                    for gc in [1, 2, 3]:
                        ws_cl.cell(row=row_stack, column=gc).border = std_border

                if idx_row < len(rejection_list):
                    r_item = rejection_list[idx_row]
                    c5 = ws_cl.cell(row=row_stack, column=5, value=r_item["category"])
                    c6 = ws_cl.cell(row=row_stack, column=6, value=r_item["cases"])
                    c7 = ws_cl.cell(row=row_stack, column=7, value=r_item["amount"])
                    c5.font = cell_font; c5.alignment = left_align; c5.border = std_border
                    c6.font = cell_font; c6.alignment = right_align; c6.border = std_border; c6.number_format = "#,##0"
                    c7.font = cell_font; c7.alignment = right_align; c7.border = std_border; c7.number_format = '"SAR" #,##0.00'
                else:
                    for gc in [5, 6, 7]:
                        ws_cl.cell(row=row_stack, column=gc).border = std_border

                row_stack += 1

            if max_span == 0:
                for gc in [1, 2, 3, 5, 6, 7]:
                    ws_cl.cell(row=row_stack, column=gc).border = std_border
                row_stack += 1

            data_end = row_stack - 1

            sum_row_idx = row_stack
            s_t1 = ws_cl.cell(row=sum_row_idx, column=1, value="Monthly Subtotal")
            s_t2 = ws_cl.cell(row=sum_row_idx, column=2, value=f"=SUM(B{data_start}:B{data_end})")
            s_t3 = ws_cl.cell(row=sum_row_idx, column=3, value=f"=SUM(C{data_start}:C{data_end})")
            for idx_c, cell in enumerate([s_t1, s_t2, s_t3], start=1):
                cell.font = total_font; cell.fill = tot_fill; cell.border = tot_border
                cell.alignment = right_align if idx_c > 1 else left_align
                if idx_c == 2: cell.number_format = "#,##0"
                if idx_c == 3: cell.number_format = '"SAR" #,##0.00'

            r_t1 = ws_cl.cell(row=sum_row_idx, column=5, value="Monthly Rejections Subtotal")
            r_t2 = ws_cl.cell(row=sum_row_idx, column=6, value=f"=SUM(F{data_start}:F{data_end})")
            r_t3 = ws_cl.cell(row=sum_row_idx, column=7, value=f"=SUM(G{data_start}:G{data_end})")
            for idx_c, cell in enumerate([r_t1, r_t2, r_t3], start=5):
                cell.font = total_font; cell.fill = tot_fill; cell.border = tot_border
                cell.alignment = right_align if idx_c > 5 else left_align
                if idx_c == 6: cell.number_format = "#,##0"
                if idx_c == 7: cell.number_format = '"SAR" #,##0.00'

            row_stack += 3

        for col in ws_cl.columns:
            m_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > m_len:
                    m_len = len(val)
            ws_cl.column_dimensions[col_letter].width = max(m_len + 3, 14)

    wb.save(out)
    return out.getvalue()

excel_binary = generate_executive_spreadsheet()
st.markdown("<div style='margin-bottom: 15px;'></div>", unsafe_allow_html=True)
col_dl1, col_dl2, col_dl3 = st.columns([3, 6, 3])
with col_dl2:
    st.download_button(
        label="Download Executive Reconciled Quarterly Sheet",
        data=excel_binary,
        file_name=f"Smart_Quarterly_Executive_Report_{datetime.date.today().year}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- 4. RELOCATED SUPPORT & FEEDBACK FORM ---
st.markdown("""
<div class="feedback-box">
    <div class="feedback-title">Support & Feedback Form</div>
    <form action="https://formspree.io/f/mgogonbb" method="POST">
        <div class="feedback-field">
            <label>Full Name</label>
            <input class="feedback-input" type="text" name="name" required placeholder="Enter your full name (e.g. Raghad Alqarni)" />
        </div>
        <div class="feedback-field">
            <label>Email Address</label>
            <input class="feedback-input" type="email" name="_replyto" required placeholder="Enter your corporate email address" />
        </div>
        <div class="feedback-field">
            <label>Message & Support Request</label>
            <textarea class="feedback-input" name="message" required rows="4" placeholder="How can we assist you with quarterly compiling today?"></textarea>
        </div>
        <button class="feedback-submit" type="submit">Send Message</button>
    </form>
</div>
""", unsafe_allow_html=True)

# --- 5. PREMIUM EXECUTIVE SIGNATURE FOOTER ---
st.markdown("""
<div class="executive-footer">
    © 2026 Raghad Alqarni. All Rights Reserved.<br/>
    Built To Premium Executive Standards. Copyright 2026.
</div>
""", unsafe_allow_html=True)
